import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="자동 견적 생성기", layout="wide")
st.title("📧 벤더 이메일 → 견적 엑셀 자동 변환기")

st.markdown("1. 📤 견적 템플릿 엑셀 업로드 (`Quotation.xlsx` 형식)")
st.markdown("2. 📋 벤더 이메일 내용을 아래에 붙여넣기")
st.markdown("3. ✅ 견적 정보 자동 추출 → 엑셀 생성")

uploaded_template = st.file_uploader("견적 템플릿 엑셀 업로드", type=["xlsx"])
vendor_text = st.text_area("벤더 이메일 / 견적 내용 붙여넣기", height=300)

if uploaded_template and vendor_text:
    # 예시: 간단한 정보 추출 (실제론 NLP 확장 가능)
    import re

    # P/N, U/P, Lead Time 추출 (간단한 정규식 기반)
    pn_match = re.search(r"P/N[:\s]+([\w\-]+)", vendor_text, re.IGNORECASE)
    price_match = re.search(r"\$?([0-9]+(?:\.[0-9]+)?)", vendor_text)
    lead_match = re.search(r"Lead Time[:\s]+([\w\s]+)", vendor_text, re.IGNORECASE)

    if pn_match and price_match:
        part_number = pn_match.group(1)
        vendor_up = float(price_match.group(1))
        lead_time = lead_match.group(1).strip() if lead_match else ""

        # 템플릿 엑셀 처리
        wb = load_workbook(uploaded_template)
        sheet = wb.active

        # 행 찾기
        target_row = None
        for row in sheet.iter_rows(min_row=8):
            cell = row[1]
            if cell.value == part_number:
                target_row = cell.row
                break

        if target_row:
            vendor_up_col = 7
            lead_time_col = 16
            margin_col = 12
            customer_up_col = 13

            margin = sheet.cell(row=target_row, column=margin_col + 1).value or 0
            customer_up = round(vendor_up * (1 + float(margin)), 2)

            sheet.cell(row=target_row, column=vendor_up_col + 1, value=vendor_up)
            sheet.cell(row=target_row, column=lead_time_col + 1, value=lead_time)
            sheet.cell(row=target_row, column=customer_up_col + 1, value=customer_up)

            # 결과 다운로드
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            st.success("견적 정보가 템플릿에 성공적으로 반영되었습니다!")
            st.download_button("📥 엑셀 다운로드", output, file_name="filled_quotation.xlsx")

        else:
            st.error("엑셀에서 해당 품번을 찾을 수 없습니다.")
    else:
        st.error("이메일에서 P/N 또는 단가를 찾을 수 없습니다.")